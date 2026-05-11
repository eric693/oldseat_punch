import json as _json
import csv
import io
from datetime import datetime as _dt, timedelta as _td, date as _date

from flask import Blueprint, request, jsonify, session

from config import TW_TZ, WEEKDAY_ZH
from db import get_db
from auth_utils import login_required, require_module
from utils import (
    shift_type_row, shift_assign_row, punch_staff_row,
    _is_holiday, _build_shift_time_map, get_schedule_config,
)

bp = Blueprint('shifts', __name__)

@bp.route('/api/shifts/types', methods=['GET'])
@require_module('sched')
def api_shift_types_list():
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM shift_types ORDER BY sort_order, id").fetchall()
    return jsonify([shift_type_row(r) for r in rows])

@bp.route('/api/shifts/types/public', methods=['GET'])
def api_shift_types_public():
    """Public endpoint for employee page."""
    with get_db() as conn:
        rows = conn.execute(
            "SELECT * FROM shift_types WHERE active=TRUE ORDER BY sort_order, id"
        ).fetchall()
    return jsonify([shift_type_row(r) for r in rows])

@bp.route('/api/shifts/types', methods=['POST'])
@require_module('sched')
def api_shift_type_create():
    b = request.get_json(force=True)
    with get_db() as conn:
        row = conn.execute("""
            INSERT INTO shift_types (name, start_time, end_time, color, departments, sort_order)
            VALUES (%s,%s,%s,%s,%s,%s) RETURNING *
        """, (b['name'], b['start_time'], b['end_time'],
              b.get('color', '#4a7bda'), b.get('departments', ''),
              int(b.get('sort_order', 0)))).fetchone()
    return jsonify(shift_type_row(row)), 201

@bp.route('/api/shifts/types/<int:sid>', methods=['PUT'])
@require_module('sched')
def api_shift_type_update(sid):
    b = request.get_json(force=True)
    with get_db() as conn:
        row = conn.execute("""
            UPDATE shift_types
            SET name=%s, start_time=%s, end_time=%s, color=%s,
                departments=%s, sort_order=%s, active=%s
            WHERE id=%s RETURNING *
        """, (b['name'], b['start_time'], b['end_time'],
              b.get('color', '#4a7bda'), b.get('departments', ''),
              int(b.get('sort_order', 0)), bool(b.get('active', True)),
              sid)).fetchone()
    return jsonify(shift_type_row(row)) if row else ('', 404)

@bp.route('/api/shifts/types/<int:sid>', methods=['DELETE'])
@require_module('sched')
def api_shift_type_delete(sid):
    with get_db() as conn:
        conn.execute("DELETE FROM shift_types WHERE id=%s", (sid,))
    return jsonify({'deleted': sid})

# ── Shift Assignments ─────────────────────────────────────────────

@bp.route('/api/shifts/assignments', methods=['GET'])
@require_module('sched')
def api_shift_assignments_list():
    month = request.args.get('month', '')
    conds, params = ['TRUE'], []
    if month:
        conds.append("to_char(sa.shift_date,'YYYY-MM')=%s"); params.append(month)
    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT sa.*,
                   ps.name as staff_name, ps.role as staff_role,
                   st.name as shift_name, st.start_time, st.end_time,
                   st.color, st.departments
            FROM shift_assignments sa
            JOIN punch_staff ps ON ps.id=sa.staff_id
            JOIN shift_types  st ON st.id=sa.shift_type_id
            WHERE {' AND '.join(conds)}
            ORDER BY sa.shift_date, ps.name
        """, params).fetchall()
    result = []
    for r in rows:
        d = shift_assign_row(r)
        d['staff_name'] = r['staff_name']
        d['staff_role'] = r['staff_role']
        d['shift_name'] = r['shift_name']
        d['start_time'] = str(r['start_time'])[:5]
        d['end_time']   = str(r['end_time'])[:5]
        d['color']      = r['color']
        result.append(d)
    return jsonify(result)


@bp.route('/api/shifts/assignments', methods=['POST'])
@require_module('sched')
def api_shift_assignment_create():
    b             = request.get_json(force=True)
    staff_ids     = b.get('staff_ids', [])
    shift_type_id = b.get('shift_type_id')
    dates         = b.get('dates', [])
    note          = b.get('note', '').strip()
    force         = bool(b.get('force', False))

    if not staff_ids or not shift_type_id or not dates:
        return jsonify({'error': '請選擇員工、班別及日期'}), 400

    created = 0
    blocked = []

    with get_db() as conn:
        # Build leave lookup for all involved staff
        leave_lookup = {}
        if not force:
            for sid in staff_ids:
                months = list({d[:7] for d in dates})
                for month in months:
                    row = conn.execute("""
                        SELECT array_agg(DISTINCT d) as dates
                        FROM schedule_requests, jsonb_array_elements_text(dates) d
                        WHERE staff_id=%s AND month=%s AND status IN ('approved','pending')
                    """, (sid, month)).fetchone()
                    if row:
                        approved_dates = row['dates'] or []
                        if isinstance(approved_dates, str):
                            try: approved_dates = _json.loads(approved_dates)
                            except: approved_dates = []
                        if sid not in leave_lookup:
                            leave_lookup[sid] = set()
                        leave_lookup[sid].update(approved_dates)

        staff_names = {}
        for r in conn.execute(
            "SELECT id,name FROM punch_staff WHERE id = ANY(%s::int[])", (staff_ids,)
        ).fetchall():
            staff_names[r['id']] = r['name']

        for sid in staff_ids:
            leave_dates = leave_lookup.get(sid, set())
            for date_str in dates:
                if date_str in leave_dates and not force:
                    blocked.append({'staff_name': staff_names.get(sid, str(sid)), 'date': date_str})
                    continue
                conn.execute("""
                    INSERT INTO shift_assignments (staff_id, shift_type_id, shift_date, note)
                    VALUES (%s,%s,%s,%s)
                    ON CONFLICT (staff_id, shift_date) DO UPDATE
                      SET shift_type_id=%s, note=%s, created_at=NOW()
                """, (sid, shift_type_id, date_str, note, shift_type_id, note))
                created += 1

    if blocked and created == 0:
        return jsonify({
            'error': '以下日期員工已有核准的排休，無法指派班別：' +
                     '、'.join([f'{x["staff_name"]} {x["date"]}' for x in blocked]),
            'blocked': blocked
        }), 422

    # Notify each assigned staff via LINE
    if created > 0:
        with get_db() as conn:
            shift_info = conn.execute(
                "SELECT name, start_time, end_time FROM shift_types WHERE id=%s", (shift_type_id,)
            ).fetchone()
        if shift_info:
            date_range = f"{min(dates)} ~ {max(dates)}" if len(dates) > 1 else dates[0]
            msg = (f"[排班通知] 已為您安排班別\n"
                   f"班別：{shift_info['name']}（{str(shift_info['start_time'])[:5]}～{str(shift_info['end_time'])[:5]}）\n"
                   f"日期：{date_range}\n"
                   f"共 {len(dates)} 天，請至員工系統查看完整排班。")
            for sid in staff_ids:
                _notify_staff_line(sid, msg)

    result = {'created': created}
    if blocked:
        result['warning'] = f'已指派 {created} 筆，跳過 {len(blocked)} 筆（員工當日有核准排休）'
        result['blocked'] = blocked
    return jsonify(result), 201


@bp.route('/api/shifts/assignments/batch-delete', methods=['POST'])
@require_module('sched')
def api_shift_assignment_batch_delete():
    b         = request.get_json(force=True)
    staff_ids = b.get('staff_ids', [])
    dates     = b.get('dates', [])
    if not staff_ids or not dates:
        return jsonify({'error': '請選擇員工及日期'}), 400
    deleted = 0
    with get_db() as conn:
        for sid in staff_ids:
            for date_str in dates:
                r = conn.execute(
                    "DELETE FROM shift_assignments WHERE staff_id=%s AND shift_date=%s RETURNING id",
                    (sid, date_str)
                ).fetchone()
                if r: deleted += 1
    return jsonify({'deleted': deleted})


@bp.route('/api/shifts/import', methods=['POST'])
@require_module('sched')
def api_shift_import():
    """
    匯入班表 CSV 或 Excel (.xlsx)。
    表頭（第一列）：姓名,日期,班別,備註  或  代碼,日期,班別,備註
    日期格式：YYYY-MM-DD
    force=1 query param 可強制覆蓋排休衝突。
    """
    import csv, io as _io
    force = request.args.get('force', '0') == '1'
    rows = []

    if 'file' in request.files:
        f = request.files['file']
        fname = (f.filename or '').lower()
        if fname.endswith('.xlsx') or fname.endswith('.xls'):
            # ── Excel 解析 ────────────────────────────────────
            import openpyxl as _opx
            wb = _opx.load_workbook(_io.BytesIO(f.read()), read_only=True, data_only=True)
            ws = wb.active
            all_rows = list(ws.values)
            if not all_rows:
                return jsonify({'error': '檔案內容為空'}), 400
            headers = [str(h).strip() if h is not None else '' for h in all_rows[0]]
            for row in all_rows[1:]:
                if all(v is None or str(v).strip() == '' for v in row):
                    continue  # skip blank rows
                d = {}
                for i, h in enumerate(headers):
                    d[h] = str(row[i]).strip() if i < len(row) and row[i] is not None else ''
                rows.append(d)
        else:
            raw = f.read().decode('utf-8-sig')
            if not raw.strip():
                return jsonify({'error': '檔案內容為空'}), 400
            reader = csv.DictReader(_io.StringIO(raw))
            if reader.fieldnames is None:
                return jsonify({'error': '無法解析 CSV 欄位'}), 400
            reader.fieldnames = [h.strip() for h in reader.fieldnames]
            rows = list(reader)
    else:
        raw = request.get_data(as_text=True)
        if not raw.strip():
            return jsonify({'error': '檔案內容為空'}), 400
        reader = csv.DictReader(_io.StringIO(raw))
        if reader.fieldnames is None:
            return jsonify({'error': '無法解析 CSV 欄位'}), 400
        reader.fieldnames = [h.strip() for h in reader.fieldnames]
        rows = list(reader)

    if not rows:
        return jsonify({'error': '無資料列'}), 400

    # 確認必要欄位
    all_keys = rows[0].keys() if rows else []
    has_name = '姓名' in all_keys
    has_code = '代碼' in all_keys
    if not (has_name or has_code):
        return jsonify({'error': '檔案缺少「姓名」或「代碼」欄位'}), 400
    if '日期' not in all_keys:
        return jsonify({'error': '檔案缺少「日期」欄位'}), 400
    if '班別' not in all_keys:
        return jsonify({'error': '檔案缺少「班別」欄位'}), 400
    with get_db() as conn:
        # 預先建立索引，避免逐列查詢
        staff_by_name = {r['name']: r['id'] for r in conn.execute(
            "SELECT id, name FROM punch_staff WHERE active=TRUE"
        ).fetchall()}
        staff_by_code = {r['employee_code']: r['id'] for r in conn.execute(
            "SELECT id, employee_code FROM punch_staff WHERE active=TRUE AND employee_code IS NOT NULL AND employee_code!=''",
        ).fetchall()}
        shift_by_name = {r['name']: r['id'] for r in conn.execute(
            "SELECT id, name FROM shift_types WHERE active=TRUE"
        ).fetchall()}

        # 預先讀取所有涉及員工的核准排休（僅在非強制時）
        leave_lookup = {}   # { staff_id: set(date_str) }
        if not force:
            leave_rows = conn.execute("""
                SELECT staff_id, dates FROM schedule_requests
                WHERE status='approved'
            """).fetchall()
            for lr in leave_rows:
                sid = lr['staff_id']
                dates_val = lr['dates']
                if isinstance(dates_val, str):
                    try: dates_val = _json.loads(dates_val)
                    except: dates_val = []
                if sid not in leave_lookup:
                    leave_lookup[sid] = set()
                leave_lookup[sid].update(dates_val or [])

        created = 0
        skipped = []   # 衝突（排休）
        errors  = []   # 找不到員工/班別、日期格式錯誤

        for i, row in enumerate(rows, start=2):   # 從第2列計算（第1列是表頭）
            name_val = row.get('姓名', '').strip()
            code_val = row.get('代碼', '').strip()
            date_str = row.get('日期', '').strip()
            shift_name = row.get('班別', '').strip()
            note = row.get('備註', '').strip()

            # 找員工 ID
            staff_id = None
            if code_val:
                staff_id = staff_by_code.get(code_val)
            if staff_id is None and name_val:
                staff_id = staff_by_name.get(name_val)
            if staff_id is None:
                errors.append({'row': i, 'reason': f'找不到員工：{code_val or name_val}'})
                continue

            # 找班別 ID
            shift_id = shift_by_name.get(shift_name)
            if shift_id is None:
                errors.append({'row': i, 'reason': f'找不到班別：{shift_name}'})
                continue

            # 驗證日期
            try:
                from datetime import date as _date
                _date.fromisoformat(date_str)
            except ValueError:
                errors.append({'row': i, 'reason': f'日期格式錯誤：{date_str}（應為 YYYY-MM-DD）'})
                continue

            # 排休衝突檢查
            if not force and date_str in leave_lookup.get(staff_id, set()):
                display = name_val or code_val
                skipped.append({'row': i, 'reason': f'{display} 於 {date_str} 有核准排休'})
                continue

            # 寫入（衝突則覆蓋）
            conn.execute("""
                INSERT INTO shift_assignments (staff_id, shift_type_id, shift_date, note)
                VALUES (%s,%s,%s,%s)
                ON CONFLICT (staff_id, shift_date) DO UPDATE
                  SET shift_type_id=%s, note=%s, created_at=NOW()
            """, (staff_id, shift_id, date_str, note, shift_id, note))
            created += 1

    result = {'created': created, 'skipped': skipped, 'errors': errors}
    if errors or skipped:
        result['message'] = f'匯入完成：{created} 筆成功，{len(skipped)} 筆排休衝突跳過，{len(errors)} 筆錯誤'
    else:
        result['message'] = f'匯入完成：共 {created} 筆排班'
    return jsonify(result), 201


@bp.route('/api/shifts/conflicts', methods=['GET'])
@require_module('sched')
def api_shift_conflicts():
    """
    偵測班表衝突與警示：
    - overtime_hours : 單班時數 > 10 小時
    - midnight_cross : 跨日班別（結束時間 < 開始時間）
    - consecutive_days : 連續排班 >= 6 天（6天警告，7天以上錯誤）
    """
    month = request.args.get('month', '')
    if not month:
        return jsonify({'error': '請指定月份'}), 400

    from datetime import date as _dc, timedelta as _tdc

    conflicts = []

    with get_db() as conn:
        rows = conn.execute("""
            SELECT sa.staff_id, sa.shift_date,
                   ps.name  AS staff_name,
                   st.name  AS shift_name,
                   st.start_time, st.end_time
            FROM shift_assignments sa
            JOIN punch_staff  ps ON ps.id = sa.staff_id
            JOIN shift_types  st ON st.id = sa.shift_type_id
            WHERE TO_CHAR(sa.shift_date, 'YYYY-MM') = %s
            ORDER BY sa.staff_id, sa.shift_date
        """, (month,)).fetchall()

    # ── 每班時數 & 跨日 ────────────────────────────────────────────
    for r in rows:
        s = r['start_time'];  e = r['end_time']
        sm = s.hour * 60 + s.minute
        em = e.hour * 60 + e.minute
        cross = em < sm
        dur   = ((24 * 60 - sm) + em) if cross else (em - sm)
        hrs   = dur / 60

        if cross:
            conflicts.append({
                'type':       'midnight_cross',
                'severity':   'info',
                'date':       str(r['shift_date']),
                'staff_name': r['staff_name'],
                'shift_name': r['shift_name'],
                'message':    f"跨日班別 {str(s)[:5]}～{str(e)[:5]}（共 {hrs:.1f} 小時）",
            })

        if hrs > 10:
            conflicts.append({
                'type':       'overtime_hours',
                'severity':   'warning' if hrs <= 12 else 'error',
                'date':       str(r['shift_date']),
                'staff_name': r['staff_name'],
                'shift_name': r['shift_name'],
                'message':    f"單班 {hrs:.1f} 小時，超過 10 小時上限",
            })

    # ── 連續排班天數 ───────────────────────────────────────────────
    staff_dates = {}
    for r in rows:
        sid = r['staff_id']
        if sid not in staff_dates:
            staff_dates[sid] = {'name': r['staff_name'], 'dates': []}
        staff_dates[sid]['dates'].append(_dc.fromisoformat(str(r['shift_date'])))

    for sid, info in staff_dates.items():
        dates = sorted(set(info['dates']))
        streak = [dates[0]]
        for i in range(1, len(dates)):
            if (dates[i] - dates[i-1]).days == 1:
                streak.append(dates[i])
            else:
                # evaluate finished streak
                if len(streak) >= 6:
                    sev = 'error' if len(streak) >= 7 else 'warning'
                    conflicts.append({
                        'type':       'consecutive_days',
                        'severity':   sev,
                        'date':       streak[0].isoformat(),
                        'staff_name': info['name'],
                        'shift_name': '',
                        'message':    (
                            f"連續排班 {len(streak)} 天"
                            f"（{streak[0].isoformat()} ～ {streak[-1].isoformat()}）"
                            + ('，違反勞基法每 7 日至少休 1 日' if len(streak) >= 7 else '，接近法定上限')
                        ),
                    })
                streak = [dates[i]]
        # last streak
        if len(streak) >= 6:
            sev = 'error' if len(streak) >= 7 else 'warning'
            conflicts.append({
                'type':       'consecutive_days',
                'severity':   sev,
                'date':       streak[0].isoformat(),
                'staff_name': info['name'],
                'shift_name': '',
                'message':    (
                    f"連續排班 {len(streak)} 天"
                    f"（{streak[0].isoformat()} ～ {streak[-1].isoformat()}）"
                    + ('，違反勞基法每 7 日至少休 1 日' if len(streak) >= 7 else '，接近法定上限')
                ),
            })

    # sort: error first, then by date
    sev_order = {'error': 0, 'warning': 1, 'info': 2}
    conflicts.sort(key=lambda c: (sev_order.get(c['severity'], 9), c['date']))
    return jsonify({'month': month, 'count': len(conflicts), 'conflicts': conflicts})


@bp.route('/api/shifts/export', methods=['GET'])
@require_module('sched')
def api_shift_export():
    """匯出指定月份班表為 Excel (.xlsx)"""
    month = request.args.get('month', '')
    if not month:
        return jsonify({'error': '請指定月份'}), 400

    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, GradientFill
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    import calendar as _cal2
    from datetime import date as _de

    y, mo = int(month[:4]), int(month[5:7])
    days_in_month = _cal2.monthrange(y, mo)[1]
    DAYS_CN = ['一','二','三','四','五','六','日']

    with get_db() as conn:
        staff_list = conn.execute(
            "SELECT id, name, employee_code, role FROM punch_staff WHERE active=TRUE ORDER BY name"
        ).fetchall()
        assigns = conn.execute("""
            SELECT sa.staff_id, sa.shift_date, sa.note,
                   st.name AS shift_name, st.start_time, st.end_time
            FROM shift_assignments sa
            JOIN shift_types st ON st.id = sa.shift_type_id
            WHERE TO_CHAR(sa.shift_date, 'YYYY-MM') = %s
        """, (month,)).fetchall()
        holidays = {str(r['date']) for r in conn.execute(
            "SELECT date FROM public_holidays WHERE TO_CHAR(date,'YYYY-MM')=%s", (month,)
        ).fetchall()}

    lookup = {}
    for a in assigns:
        key = (a['staff_id'], str(a['shift_date']))
        lookup[key] = f"{a['shift_name']}\n{str(a['start_time'])[:5]}~{str(a['end_time'])[:5]}"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{month} 班表"

    # ── 樣式定義 ─────────────────────────────────────────────
    navy_fill   = PatternFill('solid', fgColor='0F1C3A')
    grey_fill   = PatternFill('solid', fgColor='F4F6FA')
    wkend_fill  = PatternFill('solid', fgColor='FFF5F5')
    hol_fill    = PatternFill('solid', fgColor='FFF0F0')
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD'),
    )
    hdr_font  = Font(bold=True, color='FFFFFF', size=10)
    info_font = Font(bold=True, size=10)
    cell_font = Font(size=9)
    center    = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # ── 標題列 ───────────────────────────────────────────────
    ws.row_dimensions[1].height = 36
    for col, label in enumerate(['姓名', '代碼', '職稱'], start=1):
        c = ws.cell(1, col, label)
        c.font = hdr_font; c.fill = navy_fill; c.alignment = center; c.border = thin_border
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 9
    ws.column_dimensions['C'].width = 9

    for d in range(1, days_in_month + 1):
        col  = d + 3
        dt   = _de(y, mo, d)
        wd   = dt.weekday()    # 0=Mon … 6=Sun
        ds   = f"{month}-{d:02d}"
        is_wkend = wd >= 5      # Sat or Sun
        is_hol   = ds in holidays
        c = ws.cell(1, col, f"{d}\n{DAYS_CN[wd]}")
        c.font      = Font(bold=True, color='FF4444' if is_wkend or is_hol else 'FFFFFF', size=9)
        c.fill      = PatternFill('solid', fgColor='1A3060') if not (is_wkend or is_hol) else PatternFill('solid', fgColor='8B2020')
        c.alignment = center
        c.border    = thin_border
        ws.column_dimensions[get_column_letter(col)].width = 11

    # ── 員工列 ───────────────────────────────────────────────
    for row_idx, staff in enumerate(staff_list, start=2):
        ws.row_dimensions[row_idx].height = 30
        for col, val in enumerate([staff['name'], staff['employee_code'] or '', staff['role'] or ''], start=1):
            c = ws.cell(row_idx, col, val)
            c.font = info_font if col == 1 else cell_font
            c.fill = grey_fill; c.alignment = center; c.border = thin_border

        for d in range(1, days_in_month + 1):
            col = d + 3
            ds  = f"{month}-{d:02d}"
            dt  = _de(y, mo, d); wd = dt.weekday()
            val = lookup.get((staff['id'], ds), '')
            c   = ws.cell(row_idx, col, val)
            c.font      = Font(size=8, color='1A1A2E' if val else 'CCCCCC')
            c.alignment = center
            c.border    = thin_border
            if not val:
                c.fill = wkend_fill if wd >= 5 else (hol_fill if ds in holidays else PatternFill('solid', fgColor='FFFFFF'))

    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    from flask import send_file
    return send_file(
        buf, as_attachment=True,
        download_name=f"班表_{month}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@bp.route('/api/shifts/my-schedule', methods=['GET'])
def api_my_shift_schedule():
    sid = session.get('punch_staff_id')
    if not sid: return jsonify({'error': 'not logged in'}), 401
    month = request.args.get('month', '')
    with get_db() as conn:
        rows = conn.execute("""
            SELECT sa.shift_date, sa.note,
                   st.name as shift_name, st.start_time, st.end_time, st.color
            FROM shift_assignments sa
            JOIN shift_types st ON st.id=sa.shift_type_id
            WHERE sa.staff_id=%s
              AND to_char(sa.shift_date,'YYYY-MM')=%s
            ORDER BY sa.shift_date
        """, (sid, month)).fetchall()
    result = {}
    for r in rows:
        ds = r['shift_date'].isoformat()
        result[ds] = {
            'shift_name': r['shift_name'],
            'start_time': str(r['start_time'])[:5],
            'end_time':   str(r['end_time'])[:5],
            'color':      r['color'],
            'note':       r['note'],
        }
    return jsonify({'month': month, 'shifts': result})
