import csv
import io
import json as _json
from datetime import datetime as _dt, date as _date, timedelta as _td, timezone as _tz

from flask import Blueprint, request, jsonify, session, make_response

from config import TW_TZ, WEEKDAY_ZH
from db import get_db
from auth_utils import login_required, require_module
from utils import (
    punch_staff_row, punch_record_row, salary_record_row,
    leave_type_row, leave_req_row, ot_req_row,
    _build_shift_time_map, _shift_aware_day_map, _clamp_to_shift,
    _calc_annual_leave_days, _get_salary_config, _notify_review_result,
    get_schedule_config, _send_line_punch,
)

bp = Blueprint('export', __name__)

import csv
import io

@bp.route('/api/export/attendance', methods=['GET'])
@login_required
def api_export_attendance():
    """匯出月度出勤明細 CSV"""
    month    = request.args.get('month', '')
    staff_id = request.args.get('staff_id', '')
    if not month:
        from datetime import date as _de
        month = _de.today().strftime('%Y-%m')

    conds, params = ["TO_CHAR(pr.punched_at AT TIME ZONE 'Asia/Taipei','YYYY-MM')=%s"], [month]
    if staff_id:
        conds.append("pr.staff_id=%s"); params.append(int(staff_id))

    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT
                ps.employee_code,
                ps.name as staff_name,
                ps.department,
                ps.role,
                (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date as work_date,
                pr.punch_type,
                to_char(pr.punched_at AT TIME ZONE 'Asia/Taipei', 'HH24:MI') as punch_time,
                pr.is_manual,
                pr.manual_by,
                pr.gps_distance,
                pr.location_name,
                pr.note
            FROM punch_records pr
            JOIN punch_staff ps ON ps.id = pr.staff_id
            WHERE {' AND '.join(conds)}
            ORDER BY ps.name, pr.punched_at
        """, params).fetchall()

    PUNCH_LABEL = {'in':'上班打卡','out':'下班打卡','break_out':'休息開始','break_in':'休息結束'}

    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'{month} 打卡明細'
    thin = Border(left=Side(style='thin',color='DDDDDD'), right=Side(style='thin',color='DDDDDD'),
                  top=Side(style='thin',color='DDDDDD'), bottom=Side(style='thin',color='DDDDDD'))
    hfill = PatternFill('solid', fgColor='0F1C3A')
    ca = Alignment(horizontal='center', vertical='center')
    headers = ['員工代碼','姓名','部門','職稱','日期','打卡類型','時間','補打卡','操作人','GPS距離(m)','地點','備註']
    col_w   = [10, 10, 10, 12, 12, 10, 8, 6, 10, 10, 14, 16]
    for ci, (h, w) in enumerate(zip(headers, col_w), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = Font(bold=True, color='FFFFFF', name='Noto Sans TC', size=11)
        cell.fill = hfill; cell.alignment = ca; cell.border = thin
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = w
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'
    for ri, r in enumerate(rows, 2):
        vals = [r['employee_code'] or '', r['staff_name'], r['department'] or '', r['role'] or '',
                str(r['work_date']), PUNCH_LABEL.get(r['punch_type'], r['punch_type']),
                r['punch_time'], '是' if r['is_manual'] else '',
                r['manual_by'] or '',
                r['gps_distance'] if r['gps_distance'] is not None else '',
                r['location_name'] or '', r['note'] or '']
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.alignment = ca; cell.border = thin
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    from flask import Response
    return Response(buf.read(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=attendance_{month}.xlsx'})


@bp.route('/api/export/attendance-summary', methods=['GET'])
@login_required
def api_export_attendance_summary():
    """匯出月度出勤摘要 CSV（每人每天工時）"""
    month = request.args.get('month', '')
    if not month:
        from datetime import date as _df
        month = _df.today().strftime('%Y-%m')

    with get_db() as conn:
        rows = conn.execute("""
            SELECT
                pr.staff_id,
                ps.employee_code,
                ps.name,
                ps.department,
                ps.role,
                (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date as work_date,
                MIN(CASE WHEN pr.punch_type='in'  THEN to_char(pr.punched_at AT TIME ZONE 'Asia/Taipei','HH24:MI') END) as clock_in,
                MAX(CASE WHEN pr.punch_type='out' THEN to_char(pr.punched_at AT TIME ZONE 'Asia/Taipei','HH24:MI') END) as clock_out,
                MIN(CASE WHEN pr.punch_type='in'  THEN pr.punched_at AT TIME ZONE 'Asia/Taipei' END) as ci_ts,
                MAX(CASE WHEN pr.punch_type='out' THEN pr.punched_at AT TIME ZONE 'Asia/Taipei' END) as co_ts,
                BOOL_OR(pr.is_manual) as has_manual,
                COUNT(*) as punch_count
            FROM punch_records pr
            JOIN punch_staff ps ON ps.id = pr.staff_id
            WHERE TO_CHAR(pr.punched_at AT TIME ZONE 'Asia/Taipei','YYYY-MM')=%s
            GROUP BY pr.staff_id, ps.employee_code, ps.name, ps.department, ps.role,
                     (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date
            ORDER BY ps.name, (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date
        """, (month,)).fetchall()
        shift_map_xl = _build_shift_time_map(conn, month)

    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'{month} 出勤摘要'
    thin = Border(left=Side(style='thin',color='DDDDDD'), right=Side(style='thin',color='DDDDDD'),
                  top=Side(style='thin',color='DDDDDD'), bottom=Side(style='thin',color='DDDDDD'))
    hfill = PatternFill('solid', fgColor='0F1C3A')
    ca = Alignment(horizontal='center', vertical='center')
    headers = ['員工代碼','姓名','部門','職稱','日期','上班','下班','工時(h)','打卡次數','含補打']
    col_w   = [10, 10, 10, 12, 12, 8, 8, 8, 8, 6]
    for ci, (h, w) in enumerate(zip(headers, col_w), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = Font(bold=True, color='FFFFFF', name='Noto Sans TC', size=11)
        cell.fill = hfill; cell.alignment = ca; cell.border = thin
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = w
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'
    for ri, r in enumerate(rows, 2):
        dur_h = ''
        if r['ci_ts'] and r['co_ts']:
            from datetime import datetime as _dtx
            try:
                ci = r['ci_ts'] if hasattr(r['ci_ts'], 'timestamp') else _dtx.fromisoformat(str(r['ci_ts']))
                co = r['co_ts'] if hasattr(r['co_ts'], 'timestamp') else _dtx.fromisoformat(str(r['co_ts']))
                ds_xl = str(r['work_date'])
                ci, co = _clamp_to_shift(ci, co, shift_map_xl, r['staff_id'], ds_xl)
                if ci is not None:
                    gross_m = (co - ci).total_seconds() / 60
                    brk_m = 60.0 if gross_m >= 540 else (30.0 if gross_m >= 240 else 0.0)
                    dur_h = round(max(0, gross_m - brk_m) / 60, 2)
            except Exception:
                pass
        vals = [r['employee_code'] or '', r['name'], r['department'] or '', r['role'] or '',
                str(r['work_date']), r['clock_in'] or '', r['clock_out'] or '',
                dur_h, r['punch_count'], '是' if r['has_manual'] else '']
        for ci2, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci2, value=v)
            cell.alignment = ca; cell.border = thin
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    from flask import Response
    return Response(buf.read(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=attendance_summary_{month}.xlsx'})


@bp.route('/api/attendance/anomaly-report', methods=['GET'])
@login_required
def api_anomaly_report_excel():
    """匯出出勤異常報告 Excel（缺打卡、遲到、早退）"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO
    import calendar as _cal
    from datetime import datetime as _dtx, timedelta as _tdx

    month = request.args.get('month', '') or _dt.now(TW_TZ).strftime('%Y-%m')
    try:
        y, mo = int(month[:4]), int(month[5:7])
    except Exception:
        return jsonify({'error': '月份格式錯誤'}), 400

    TW_OFF = _tdx(hours=8)

    with get_db() as conn:
        punch_rows = conn.execute("""
            SELECT ps.id as staff_id, ps.name as staff_name,
                   ps.department,
                   (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date as work_date,
                   MIN(CASE WHEN pr.punch_type='in'  THEN (pr.punched_at AT TIME ZONE 'Asia/Taipei') END) as clock_in,
                   MAX(CASE WHEN pr.punch_type='out' THEN (pr.punched_at AT TIME ZONE 'Asia/Taipei') END) as clock_out,
                   BOOL_OR(pr.punch_type='in')  as has_in,
                   BOOL_OR(pr.punch_type='out') as has_out
            FROM punch_records pr
            JOIN punch_staff ps ON ps.id=pr.staff_id AND ps.active=TRUE
            WHERE TO_CHAR(pr.punched_at AT TIME ZONE 'Asia/Taipei','YYYY-MM')=%s
            GROUP BY ps.id, ps.name, ps.department,
                     (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date
            ORDER BY work_date, ps.name
        """, (month,)).fetchall()

        shift_rows = conn.execute("""
            SELECT sa.staff_id, sa.shift_date,
                   st.start_time::text as start_time,
                   st.end_time::text   as end_time
            FROM shift_assignments sa
            JOIN shift_types st ON st.id=sa.shift_type_id
            WHERE TO_CHAR(sa.shift_date,'YYYY-MM')=%s
        """, (month,)).fetchall()

        y_int = int(month[:4]); mo_int = int(month[5:7])
        first_day = f"{y_int}-{mo_int:02d}-01"
        days_in   = _cal.monthrange(y_int, mo_int)[1]
        last_day  = f"{y_int}-{mo_int:02d}-{days_in:02d}"
        leave_rows = conn.execute("""
            SELECT staff_id, start_date, end_date
            FROM leave_requests
            WHERE status='approved'
              AND start_date <= %s AND end_date >= %s
        """, (last_day, first_day)).fetchall()

    # Build lookup maps
    shift_map = {(r['staff_id'], str(r['shift_date'])): r for r in shift_rows}
    leave_set = set()
    from datetime import date as _dax, timedelta as _tdax
    for lr in leave_rows:
        s = lr['start_date']; e = lr['end_date']
        cur = s
        while cur <= e:
            leave_set.add((lr['staff_id'], str(cur)))
            cur = _dax.fromisoformat(str(cur)) + _tdax(days=1)
            cur = cur if isinstance(cur, _dax) else cur.date()

    today = _dax.today()

    # Build anomaly rows
    anomalies = []
    for r in punch_rows:
        ds = str(r['work_date'])
        sid = r['staff_id']
        shift = shift_map.get((sid, ds))

        anomaly_type = ''; detail = ''
        late_min = 0; early_min = 0

        if not r['has_in'] and r['has_out']:
            anomaly_type = '缺上班打卡'; detail = f"僅有下班 {str(r['clock_out'])[11:16]}"
        elif r['has_in'] and not r['has_out']:
            if _dax.fromisoformat(ds) < today:
                anomaly_type = '缺下班打卡'; detail = f"上班 {str(r['clock_in'])[11:16]} 無下班"
        elif r['has_in'] and r['has_out'] and shift:
            ci_t = str(r['clock_in'])[11:16]
            co_t = str(r['clock_out'])[11:16]
            sh_s = str(shift['start_time'])[:5]
            sh_e = str(shift['end_time'])[:5]
            try:
                ci_m = int(ci_t[:2])*60 + int(ci_t[3:5])
                sh_s_m = int(sh_s[:2])*60 + int(sh_s[3:5])
                if ci_m - sh_s_m > 10:
                    late_min = ci_m - sh_s_m
                    anomaly_type = '遲到'; detail = f"應 {sh_s}，實際 {ci_t}（+{late_min}分）"
            except Exception:
                pass
            if not anomaly_type:
                try:
                    co_m = int(co_t[:2])*60 + int(co_t[3:5])
                    sh_e_m = int(sh_e[:2])*60 + int(sh_e[3:5])
                    if sh_e_m - co_m > 15:
                        early_min = sh_e_m - co_m
                        anomaly_type = '早退'; detail = f"應 {sh_e}，實際 {co_t}（-{early_min}分）"
                except Exception:
                    pass

        if anomaly_type:
            anomalies.append({
                'staff_name':  r['staff_name'],
                'department':  r['department'] or '',
                'date':        ds,
                'shift_start': str(shift['start_time'])[:5] if shift else '—',
                'shift_end':   str(shift['end_time'])[:5]   if shift else '—',
                'clock_in':    str(r['clock_in'])[11:16]  if r['clock_in']  else '—',
                'clock_out':   str(r['clock_out'])[11:16] if r['clock_out'] else '—',
                'anomaly_type': anomaly_type,
                'detail':       detail,
            })

    # Build Excel
    wb   = openpyxl.Workbook()
    ws   = wb.active
    ws.title = f'{month} 異常明細'

    thin = Border(
        left=Side(style='thin', color='DDDDDD'), right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin',  color='DDDDDD'), bottom=Side(style='thin', color='DDDDDD'),
    )
    header_fill   = PatternFill('solid', fgColor='0F1C3A')
    warn_fill     = PatternFill('solid', fgColor='FFF3CD')
    err_fill      = PatternFill('solid', fgColor='FDECEA')
    center_align  = Alignment(horizontal='center', vertical='center')

    headers = ['員工姓名','部門','日期','應上班','應下班','實際上班','實際下班','異常類型','說明']
    col_w   = [12, 10, 12, 8, 8, 8, 8, 12, 30]
    for ci, (h, w) in enumerate(zip(headers, col_w), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = Font(bold=True, color='FFFFFF', name='Noto Sans TC', size=11)
        cell.fill      = header_fill
        cell.alignment = center_align
        cell.border    = thin
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = w

    for ri, a in enumerate(anomalies, 2):
        row_fill = err_fill if a['anomaly_type'] in ('缺上班打卡','缺下班打卡') else warn_fill
        vals = [a['staff_name'], a['department'], a['date'],
                a['shift_start'], a['shift_end'],
                a['clock_in'], a['clock_out'],
                a['anomaly_type'], a['detail']]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.fill      = row_fill
            cell.alignment = center_align if ci != 9 else Alignment(vertical='center')
            cell.border    = thin

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'

    # Summary sheet
    ws2 = wb.create_sheet('摘要')
    ws2.append(['統計', '數量'])
    ws2.append(['異常總筆數', len(anomalies)])
    by_type = {}
    for a in anomalies:
        by_type[a['anomaly_type']] = by_type.get(a['anomaly_type'], 0) + 1
    for t, c in sorted(by_type.items(), key=lambda x: -x[1]):
        ws2.append([t, c])

    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    from flask import Response as _FR
    return _FR(
        buf.read(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=anomaly_{month}.xlsx'}
    )


@bp.route('/api/export/salary', methods=['GET'])
@login_required
def api_export_salary():
    """匯出月度薪資明細 CSV"""
    month = request.args.get('month', '')
    if not month:
        from datetime import date as _dg
        month = _dg.today().strftime('%Y-%m')

    with get_db() as conn:
        rows = conn.execute("""
            SELECT sr.*, ps.name as staff_name, ps.employee_code,
                   ps.department, ps.role, ps.salary_type
            FROM salary_records sr
            JOIN punch_staff ps ON ps.id = sr.staff_id
            WHERE sr.month = %s
            ORDER BY ps.name
        """, (month,)).fetchall()

    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'{month} 薪資明細'
    thin = Border(left=Side(style='thin',color='DDDDDD'), right=Side(style='thin',color='DDDDDD'),
                  top=Side(style='thin',color='DDDDDD'), bottom=Side(style='thin',color='DDDDDD'))
    hfill = PatternFill('solid', fgColor='0F1C3A')
    ca = Alignment(horizontal='center', vertical='center')
    headers = ['員工代碼','姓名','部門','職稱','薪資制度',
               '工作日','出勤天數','請假天數','無薪假天數',
               '津貼合計','扣除合計','加班費','實領金額','發薪日','狀態','備註']
    col_w = [10,10,10,12,8,8,8,8,8,10,10,10,12,12,8,16]
    for ci, (h, w) in enumerate(zip(headers, col_w), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = Font(bold=True, color='FFFFFF', name='Noto Sans TC', size=11)
        cell.fill = hfill; cell.alignment = ca; cell.border = thin
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = w
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'
    for ri, r in enumerate(rows, 2):
        sal_type = r['salary_type'] or 'monthly'
        pay_date_val = r['pay_date'].isoformat() if r['pay_date'] else ''
        vals = [r['employee_code'] or '', r['staff_name'], r['department'] or '', r['role'] or '',
                '時薪制' if sal_type == 'hourly' else '月薪制',
                float(r['work_days'] or 0), float(r['actual_days'] or 0),
                float(r['leave_days'] or 0), float(r['unpaid_days'] or 0),
                float(r['allowance_total'] or 0), float(r['deduction_total'] or 0),
                float(r['ot_pay'] or 0), float(r['net_pay'] or 0),
                pay_date_val, '已確認' if r['status'] == 'confirmed' else '草稿', r['note'] or '']
        for ci2, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci2, value=v)
            cell.alignment = ca; cell.border = thin
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    from flask import Response
    return Response(buf.read(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=salary_{month}.xlsx'})


@bp.route('/api/export/leave', methods=['GET'])
@login_required
def api_export_leave():
    """匯出請假記錄 CSV"""
    month    = request.args.get('month', '')
    year     = request.args.get('year',  '')
    staff_id = request.args.get('staff_id', '')

    conds, params = ['lr.status=%s'], ['approved']
    if month: conds.append("to_char(lr.start_date,'YYYY-MM')=%s"); params.append(month)
    if year:  conds.append("EXTRACT(YEAR FROM lr.start_date)=%s"); params.append(int(year))
    if staff_id: conds.append("lr.staff_id=%s"); params.append(int(staff_id))

    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT lr.*, ps.name as staff_name, ps.employee_code,
                   ps.department, lt.name as leave_type_name, lt.pay_rate
            FROM leave_requests lr
            JOIN punch_staff ps ON ps.id = lr.staff_id
            JOIN leave_types  lt ON lt.id = lr.leave_type_id
            WHERE {' AND '.join(conds)}
            ORDER BY lr.start_date, ps.name
        """, params).fetchall()

    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '請假記錄'
    thin = Border(left=Side(style='thin',color='DDDDDD'), right=Side(style='thin',color='DDDDDD'),
                  top=Side(style='thin',color='DDDDDD'), bottom=Side(style='thin',color='DDDDDD'))
    hfill = PatternFill('solid', fgColor='0F1C3A')
    ca = Alignment(horizontal='center', vertical='center')
    headers = ['員工代碼','姓名','部門','假別','薪資倍率','開始日期','結束日期','天數','原因','代理人','狀態']
    col_w = [10,10,10,10,8,12,12,6,20,10,8]
    for ci, (h, w) in enumerate(zip(headers, col_w), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = Font(bold=True, color='FFFFFF', name='Noto Sans TC', size=11)
        cell.fill = hfill; cell.alignment = ca; cell.border = thin
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = w
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'
    PAY_LABEL = {1.0:'全薪', 0.5:'半薪', 0.0:'無薪'}
    STATUS_LABEL = {'approved':'已核准','rejected':'已退回','pending':'待審核'}
    for ri, r in enumerate(rows, 2):
        vals = [r['employee_code'] or '', r['staff_name'], r['department'] or '',
                r['leave_type_name'], PAY_LABEL.get(float(r['pay_rate']), f"{r['pay_rate']}倍"),
                str(r['start_date']), str(r['end_date']), float(r['total_days']),
                r['reason'] or '', r['substitute_name'] or '',
                STATUS_LABEL.get(r['status'], r['status'])]
        for ci2, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci2, value=v)
            cell.alignment = Alignment(horizontal='center', vertical='center') if ci2 != 9 else Alignment(vertical='center')
            cell.border = thin
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    from flask import Response
    return Response(buf.read(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=leave_{month or year or "all"}.xlsx'})


@bp.route('/api/export/overtime', methods=['GET'])
@login_required
def api_export_overtime():
    """匯出加班申請記錄 Excel"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO
    month  = request.args.get('month', '')
    status = request.args.get('status', '')
    conds, params = ['TRUE'], []
    if month:  conds.append("to_char(r.request_date,'YYYY-MM')=%s"); params.append(month)
    if status: conds.append("r.status=%s"); params.append(status)
    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT r.*, ps.name as staff_name, ps.employee_code, ps.department, ps.role as staff_role
            FROM overtime_requests r
            JOIN punch_staff ps ON ps.id=r.staff_id
            WHERE {' AND '.join(conds)}
            ORDER BY r.request_date DESC, r.created_at DESC
        """, params).fetchall()
    wb = openpyxl.Workbook(); ws = wb.active
    ws.title = f'{month or "全部"} 加班記錄'
    thin = Border(left=Side(style='thin',color='DDDDDD'), right=Side(style='thin',color='DDDDDD'),
                  top=Side(style='thin',color='DDDDDD'), bottom=Side(style='thin',color='DDDDDD'))
    hfill = PatternFill('solid', fgColor='0F1C3A')
    ca = Alignment(horizontal='center', vertical='center')
    headers = ['員工代碼','姓名','部門','職稱','申請日期','開始時間','結束時間','加班時數','日期類型','加班費','狀態','審核人','審核意見','申請原因']
    col_w   = [10,10,10,12,12,8,8,8,10,10,8,10,16,20]
    for ci, (h, w) in enumerate(zip(headers, col_w), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = Font(bold=True, color='FFFFFF', name='Noto Sans TC', size=11)
        cell.fill = hfill; cell.alignment = ca; cell.border = thin
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = w
    ws.row_dimensions[1].height = 22; ws.freeze_panes = 'A2'
    DAY_TYPE = {'weekday':'平日','rest_day':'休息日','holiday':'國定假日','special':'例假日'}
    STATUS_L  = {'pending':'待審核','approved':'已核准','rejected':'已退回'}
    approved_fill = PatternFill('solid', fgColor='E8F5E9')
    rejected_fill = PatternFill('solid', fgColor='FDECEA')
    for ri, r in enumerate(rows, 2):
        vals = [r['employee_code'] or '', r['staff_name'], r['department'] or '', r['staff_role'] or '',
                str(r['request_date']),
                str(r['start_time'])[:5] if r.get('start_time') else '',
                str(r['end_time'])[:5]   if r.get('end_time')   else '',
                float(r['ot_hours'] or 0),
                DAY_TYPE.get(r.get('day_type','weekday'),'平日'),
                float(r['ot_pay'] or 0),
                STATUS_L.get(r['status'], r['status']),
                r['reviewed_by'] or '', r['review_note'] or '', r.get('reason','') or '']
        row_fill = (approved_fill if r['status']=='approved' else
                    rejected_fill if r['status']=='rejected' else None)
        for ci2, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci2, value=v)
            if row_fill: cell.fill = row_fill
            cell.alignment = ca if ci2 not in (13,14) else Alignment(vertical='center')
            cell.border = thin
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    from flask import Response
    return Response(buf.read(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=overtime_{month or "all"}.xlsx'})


@bp.route('/api/export/training', methods=['GET'])
@login_required
def api_export_training():
    """匯出教育訓練記錄 Excel"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO
    staff_id = request.args.get('staff_id', '')
    category = request.args.get('category', '')
    conds, params = ['TRUE'], []
    if staff_id: conds.append("tr.staff_id=%s"); params.append(int(staff_id))
    if category: conds.append("tr.category=%s"); params.append(category)
    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT tr.*, ps.name AS staff_name, ps.employee_code, ps.department
            FROM training_records tr
            JOIN punch_staff ps ON tr.staff_id = ps.id
            WHERE {' AND '.join(conds)}
            ORDER BY tr.expiry_date ASC NULLS LAST, tr.completed_date DESC
        """, params).fetchall()
    from datetime import date as _da
    today = _da.today()
    wb = openpyxl.Workbook(); ws = wb.active
    ws.title = '教育訓練記錄'
    thin = Border(left=Side(style='thin',color='DDDDDD'), right=Side(style='thin',color='DDDDDD'),
                  top=Side(style='thin',color='DDDDDD'), bottom=Side(style='thin',color='DDDDDD'))
    hfill = PatternFill('solid', fgColor='0F1C3A')
    ca = Alignment(horizontal='center', vertical='center')
    headers = ['員工代碼','姓名','部門','課程名稱','類別','完成日期','到期日期','剩餘天數','狀態','證書號碼','備註']
    col_w   = [10,10,10,20,10,12,12,8,8,14,16]
    for ci, (h, w) in enumerate(zip(headers, col_w), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = Font(bold=True, color='FFFFFF', name='Noto Sans TC', size=11)
        cell.fill = hfill; cell.alignment = ca; cell.border = thin
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = w
    ws.row_dimensions[1].height = 22; ws.freeze_panes = 'A2'
    CAT_L = {'food_safety':'食品安全','fire':'消防訓練','first_aid':'急救訓練','hygiene':'衛生訓練','general':'一般訓練'}
    valid_fill    = PatternFill('solid', fgColor='E8F5E9')
    expiring_fill = PatternFill('solid', fgColor='FFF3CD')
    expired_fill  = PatternFill('solid', fgColor='FDECEA')
    for ri, r in enumerate(rows, 2):
        days_left = ''; status_label = '無到期'
        if r['expiry_date']:
            ed = _da.fromisoformat(str(r['expiry_date']))
            days_left = (ed - today).days
            status_label = '已過期' if days_left < 0 else ('即將到期' if days_left <= 60 else '有效')
        row_fill = (expired_fill if status_label=='已過期' else
                    expiring_fill if status_label=='即將到期' else valid_fill)
        vals = [r['employee_code'] or '', r['staff_name'], r['department'] or '',
                r['course_name'], CAT_L.get(r['category'], r['category'] or ''),
                str(r['completed_date']) if r['completed_date'] else '',
                str(r['expiry_date'])    if r['expiry_date']    else '',
                days_left if isinstance(days_left, int) else '',
                status_label, r['certificate_no'] or '', r['note'] or '']
        for ci2, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci2, value=v)
            cell.fill = row_fill
            cell.alignment = ca if ci2 not in (4,11) else Alignment(vertical='center')
            cell.border = thin
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    from flask import Response
    return Response(buf.read(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=training_records.xlsx'})


@bp.route('/api/export/performance', methods=['GET'])
@login_required
def api_export_performance():
    """匯出績效考核記錄 Excel"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO
    staff_id = request.args.get('staff_id', '')
    period   = request.args.get('period', '')
    conds, params = ['TRUE'], []
    if staff_id: conds.append("pr.staff_id=%s"); params.append(int(staff_id))
    if period:   conds.append("pr.period_label ILIKE %s"); params.append(f'%{period}%')
    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT pr.*, ps.name AS staff_name, ps.employee_code, ps.department,
                   ps.role AS staff_role, pt.name AS tpl_name
            FROM performance_reviews pr
            JOIN punch_staff ps ON ps.id = pr.staff_id
            LEFT JOIN performance_templates pt ON pt.id = pr.template_id
            WHERE {' AND '.join(conds)}
            ORDER BY pr.reviewed_at DESC
        """, params).fetchall()
    wb = openpyxl.Workbook(); ws = wb.active
    ws.title = '績效考核記錄'
    thin = Border(left=Side(style='thin',color='DDDDDD'), right=Side(style='thin',color='DDDDDD'),
                  top=Side(style='thin',color='DDDDDD'), bottom=Side(style='thin',color='DDDDDD'))
    hfill = PatternFill('solid', fgColor='0F1C3A')
    ca = Alignment(horizontal='center', vertical='center')
    headers = ['員工代碼','姓名','部門','職稱','考核期間','範本','總分','滿分','百分比','等級','備註','考核人','考核日期']
    col_w   = [10,10,10,12,14,14,8,8,8,6,20,10,14]
    for ci, (h, w) in enumerate(zip(headers, col_w), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = Font(bold=True, color='FFFFFF', name='Noto Sans TC', size=11)
        cell.fill = hfill; cell.alignment = ca; cell.border = thin
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = w
    ws.row_dimensions[1].height = 22; ws.freeze_panes = 'A2'
    GRADE_FILL = {'S': PatternFill('solid', fgColor='FFD700'),
                  'A': PatternFill('solid', fgColor='E8F5E9'),
                  'B': PatternFill('solid', fgColor='E3F2FD'),
                  'C': PatternFill('solid', fgColor='FFF3CD'),
                  'D': PatternFill('solid', fgColor='FDECEA')}
    for ri, r in enumerate(rows, 2):
        total = float(r['total_score'] or 0); max_s = float(r['max_score'] or 100)
        pct = round(total / max_s * 100, 1) if max_s > 0 else 0
        reviewed_at = str(r['reviewed_at'])[:10] if r['reviewed_at'] else ''
        vals = [r['employee_code'] or '', r['staff_name'], r['department'] or '',
                r['staff_role'] or '', r['period_label'] or '',
                r['tpl_name'] or '', total, max_s, pct,
                r['grade'] or '', r['comments'] or '', r['reviewer'] or '', reviewed_at]
        grade = r['grade'] or ''
        row_fill = GRADE_FILL.get(grade)
        for ci2, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci2, value=v)
            if row_fill: cell.fill = row_fill
            cell.alignment = ca if ci2 != 11 else Alignment(vertical='center')
            cell.border = thin
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    from flask import Response
    return Response(buf.read(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=performance_reviews.xlsx'})


@bp.route('/api/export/expense', methods=['GET'])
@login_required
def api_export_expense():
    """匯出費用報帳記錄 Excel"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO
    status = request.args.get('status', '')
    month  = request.args.get('month', '')
    conds, params = ['TRUE'], []
    if status: conds.append("ec.status=%s"); params.append(status)
    if month:  conds.append("to_char(ec.expense_date,'YYYY-MM')=%s"); params.append(month)
    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT ec.*, ps.name as staff_name, ps.employee_code, ps.department
            FROM expense_claims ec
            JOIN punch_staff ps ON ps.id=ec.staff_id
            WHERE {' AND '.join(conds)}
            ORDER BY ec.expense_date DESC, ec.created_at DESC
        """, params).fetchall()
    wb = openpyxl.Workbook(); ws = wb.active
    ws.title = f'{month or "全部"} 費用報帳'
    thin = Border(left=Side(style='thin',color='DDDDDD'), right=Side(style='thin',color='DDDDDD'),
                  top=Side(style='thin',color='DDDDDD'), bottom=Side(style='thin',color='DDDDDD'))
    hfill = PatternFill('solid', fgColor='0F1C3A')
    ca = Alignment(horizontal='center', vertical='center')
    headers = ['員工代碼','姓名','部門','申請日期','類型','標題','金額','說明','狀態','審核人','審核意見','審核日期']
    col_w   = [10,10,10,12,10,20,10,20,8,10,16,14]
    for ci, (h, w) in enumerate(zip(headers, col_w), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = Font(bold=True, color='FFFFFF', name='Noto Sans TC', size=11)
        cell.fill = hfill; cell.alignment = ca; cell.border = thin
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = w
    ws.row_dimensions[1].height = 22; ws.freeze_panes = 'A2'
    STATUS_L = {'pending':'待審核','approved':'已核准','rejected':'已拒絕'}
    approved_fill = PatternFill('solid', fgColor='E8F5E9')
    rejected_fill = PatternFill('solid', fgColor='FDECEA')
    for ri, r in enumerate(rows, 2):
        reviewed_at = str(r['reviewed_at'])[:10] if r.get('reviewed_at') else ''
        vals = [r['employee_code'] or '', r['staff_name'], r['department'] or '',
                str(r['expense_date']) if r.get('expense_date') else '',
                r.get('category') or '', r['title'] or '',
                float(r['amount'] or 0), r['note'] or '',
                STATUS_L.get(r['status'], r['status']),
                r['reviewed_by'] or '', r['review_note'] or '', reviewed_at]
        row_fill = (approved_fill if r['status']=='approved' else
                    rejected_fill if r['status']=='rejected' else None)
        for ci2, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci2, value=v)
            if row_fill: cell.fill = row_fill
            cell.alignment = ca if ci2 not in (8,11) else Alignment(vertical='center')
            cell.border = thin
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    from flask import Response
    return Response(buf.read(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=expense_{month or "all"}.xlsx'})


@bp.route('/api/export/staff', methods=['GET'])
@login_required
def api_export_staff():
    """匯出員工名單 Excel"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO
    include_inactive = request.args.get('include_inactive', '') == '1'
    cond = '' if include_inactive else 'WHERE active=TRUE'
    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT employee_code, name, department, position_title, role,
                   salary_type, base_salary, hourly_rate, daily_hours,
                   hire_date, birth_date, national_id, active
            FROM punch_staff {cond}
            ORDER BY department, name
        """).fetchall()
    wb = openpyxl.Workbook(); ws = wb.active
    ws.title = '員工名單'
    thin = Border(left=Side(style='thin',color='DDDDDD'), right=Side(style='thin',color='DDDDDD'),
                  top=Side(style='thin',color='DDDDDD'), bottom=Side(style='thin',color='DDDDDD'))
    hfill = PatternFill('solid', fgColor='0F1C3A')
    ca = Alignment(horizontal='center', vertical='center')
    headers = ['員工代碼','姓名','部門','職位','職稱','薪資制度','底薪','時薪','每日工時','到職日','生日','身分證','狀態']
    col_w   = [10,10,10,12,12,8,10,10,8,12,12,14,6]
    for ci, (h, w) in enumerate(zip(headers, col_w), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = Font(bold=True, color='FFFFFF', name='Noto Sans TC', size=11)
        cell.fill = hfill; cell.alignment = ca; cell.border = thin
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = w
    ws.row_dimensions[1].height = 22; ws.freeze_panes = 'A2'
    inactive_fill = PatternFill('solid', fgColor='F5F5F5')
    for ri, r in enumerate(rows, 2):
        vals = [r['employee_code'] or '', r['name'], r['department'] or '',
                r['position_title'] or '', r['role'] or '',
                '時薪制' if r['salary_type']=='hourly' else '月薪制',
                float(r['base_salary'] or 0), float(r['hourly_rate'] or 0),
                float(r['daily_hours'] or 8),
                str(r['hire_date']) if r['hire_date'] else '',
                str(r['birth_date']) if r['birth_date'] else '',
                r['national_id'] or '',
                '在職' if r['active'] else '離職']
        for ci2, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci2, value=v)
            if not r['active']: cell.fill = inactive_fill
            cell.alignment = ca; cell.border = thin
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    from flask import Response
    return Response(buf.read(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=staff_list.xlsx'})



@bp.route('/api/dashboard', methods=['GET'])
@login_required
def api_dashboard():
    from datetime import date as _dd, datetime as _ddt, timezone as _tz, timedelta as _tdd
    TW    = _tz(_tdd(hours=8))
    today = _ddt.now(TW).date()

    # 支援傳入月份參數；預設為當月
    req_month = request.args.get('month', '').strip()
    if req_month and len(req_month) == 7:
        month = req_month
        try:
            y, m = int(month[:4]), int(month[5:])
            import calendar as _cal_d
            last_day = _cal_d.monthrange(y, m)[1]
            from datetime import date as _dcheck
            # 如果查詢的是未來月份，today 仍用實際今天
        except Exception:
            month = today.strftime('%Y-%m')
    else:
        month = today.strftime('%Y-%m')

    with get_db() as conn:

        # ── 今日出勤狀況 ─────────────────────────────────────────
        total_staff = conn.execute(
            "SELECT COUNT(*) as c FROM punch_staff WHERE active=TRUE"
        ).fetchone()['c']

        # 今日已打上班卡的人數
        clocked_in = conn.execute("""
            SELECT COUNT(DISTINCT staff_id) as c
            FROM punch_records
            WHERE punch_type='in'
              AND (punched_at AT TIME ZONE 'Asia/Taipei')::date = %s
        """, (today,)).fetchone()['c']

        # 今日已打下班卡的人數
        clocked_out = conn.execute("""
            SELECT COUNT(DISTINCT staff_id) as c
            FROM punch_records
            WHERE punch_type='out'
              AND (punched_at AT TIME ZONE 'Asia/Taipei')::date = %s
        """, (today,)).fetchone()['c']

        # 今日請假人數（已核准）
        on_leave_today = conn.execute("""
            SELECT COUNT(DISTINCT staff_id) as c
            FROM leave_requests
            WHERE status='approved'
              AND start_date <= %s AND end_date >= %s
        """, (today, today)).fetchone()['c']

        # 今日出勤明細（每人狀態）
        today_detail_rows = conn.execute("""
            SELECT ps.id, ps.name, ps.role,
                   MAX(CASE WHEN pr.punch_type='in'  THEN to_char(pr.punched_at AT TIME ZONE 'Asia/Taipei','HH24:MI') END) as clock_in,
                   MAX(CASE WHEN pr.punch_type='out' THEN to_char(pr.punched_at AT TIME ZONE 'Asia/Taipei','HH24:MI') END) as clock_out,
                   COUNT(pr.id) as punch_count
            FROM punch_staff ps
            LEFT JOIN punch_records pr
              ON pr.staff_id = ps.id
              AND (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date = %s
            WHERE ps.active = TRUE
            GROUP BY ps.id, ps.name, ps.role
            ORDER BY ps.name
        """, (today,)).fetchall()

        today_detail = []
        for r in today_detail_rows:
            # Check if on leave
            leave_row = conn.execute("""
                SELECT lt.name as leave_name
                FROM leave_requests lr
                JOIN leave_types lt ON lt.id = lr.leave_type_id
                WHERE lr.staff_id=%s AND lr.status='approved'
                  AND lr.start_date <= %s AND lr.end_date >= %s
                LIMIT 1
            """, (r['id'], today, today)).fetchone()

            if r['clock_in']:
                if r['clock_out']:
                    status = 'done'
                    status_label = '已下班'
                else:
                    status = 'working'
                    status_label = '上班中'
            elif leave_row:
                status = 'leave'
                status_label = leave_row['leave_name']
            else:
                status = 'absent'
                status_label = '未出勤'

            today_detail.append({
                'id':           r['id'],
                'name':         r['name'],
                'role':         r['role'] or '',
                'clock_in':     r['clock_in']  or '',
                'clock_out':    r['clock_out'] or '',
                'punch_count':  r['punch_count'],
                'status':       status,
                'status_label': status_label,
            })

        # ── 待審申請數 ───────────────────────────────────────────
        pending_punch   = conn.execute("SELECT COUNT(*) as c FROM punch_requests WHERE status='pending'").fetchone()['c']
        pending_ot      = conn.execute("SELECT COUNT(*) as c FROM overtime_requests WHERE status='pending'").fetchone()['c']
        pending_sched   = conn.execute("SELECT COUNT(*) as c FROM schedule_requests WHERE status IN ('pending','modified_pending')").fetchone()['c']
        pending_leave   = conn.execute("SELECT COUNT(*) as c FROM leave_requests WHERE status='pending'").fetchone()['c']

        # ── 本月薪資總覽 ─────────────────────────────────────────
        sal_rows = conn.execute("""
            SELECT COUNT(*) as total_count,
                   COUNT(*) FILTER (WHERE status='confirmed') as confirmed_count,
                   COALESCE(SUM(net_pay),0) as total_net,
                   COALESCE(SUM(allowance_total),0) as total_allow,
                   COALESCE(SUM(deduction_total),0) as total_deduct
            FROM salary_records WHERE month=%s
        """, (month,)).fetchone()

        # ── 本月出勤統計（每天出勤人數，用於折線圖）─────────────
        import calendar as _cal
        days_in_month = _cal.monthrange(today.year, today.month)[1]
        daily_rows = conn.execute("""
            SELECT (punched_at AT TIME ZONE 'Asia/Taipei')::date as d,
                   COUNT(DISTINCT staff_id) as cnt
            FROM punch_records
            WHERE punch_type='in'
              AND to_char(punched_at AT TIME ZONE 'Asia/Taipei','YYYY-MM')=%s
            GROUP BY (punched_at AT TIME ZONE 'Asia/Taipei')::date
            ORDER BY d
        """, (month,)).fetchall()
        daily_map = {str(r['d']): r['cnt'] for r in daily_rows}
        daily_attendance = []
        for day in range(1, days_in_month + 1):
            ds = f"{month}-{day:02d}"
            dt = _dd(today.year, today.month, day)
            daily_attendance.append({
                'date':    ds,
                'day':     day,
                'count':   daily_map.get(ds, 0),
                'is_past': dt <= today,
                'weekday': dt.weekday(),
            })

        # ── 本月請假類型分佈（圓餅圖）───────────────────────────
        leave_dist_rows = conn.execute("""
            SELECT lt.name, lt.color, COUNT(*) as cnt,
                   COALESCE(SUM(lr.total_days),0) as days
            FROM leave_requests lr
            JOIN leave_types lt ON lt.id = lr.leave_type_id
            WHERE lr.status='approved'
              AND to_char(lr.start_date,'YYYY-MM')=%s
            GROUP BY lt.name, lt.color
            ORDER BY days DESC
        """, (month,)).fetchall()
        leave_distribution = [
            {'name': r['name'], 'color': r['color'], 'count': r['cnt'], 'days': float(r['days'])}
            for r in leave_dist_rows
        ]

        # ── 本月加班費排行（橫條圖）─────────────────────────────
        ot_rank_rows = conn.execute("""
            SELECT ps.name, ps.role,
                   COALESCE(SUM(r.ot_pay),0) as total_pay,
                   COALESCE(SUM(r.ot_hours),0) as total_hours
            FROM overtime_requests r
            JOIN punch_staff ps ON ps.id = r.staff_id
            WHERE r.status='approved'
              AND to_char(r.request_date,'YYYY-MM')=%s
            GROUP BY ps.name, ps.role
            ORDER BY total_pay DESC
            LIMIT 8
        """, (month,)).fetchall()
        ot_ranking = [
            {'name': r['name'], 'role': r['role'] or '', 'pay': float(r['total_pay']), 'hours': float(r['total_hours'])}
            for r in ot_rank_rows
        ]

    from datetime import date as _ddc
    cur_month = _ddc.today().strftime('%Y-%m')
    return jsonify({
        'month':            month,
        'today':            str(today),
        'is_current_month': month == cur_month,
        # 今日出勤
        'today_summary': {
            'total':       total_staff,
            'working':     clocked_in - clocked_out,
            'clocked_in':  clocked_in,
            'clocked_out': clocked_out,
            'on_leave':    on_leave_today,
            'absent':      total_staff - clocked_in - on_leave_today,
        },
        'today_detail': today_detail,
        # 待審申請
        'pending': {
            'punch':  pending_punch,
            'ot':     pending_ot,
            'sched':  pending_sched,
            'leave':  pending_leave,
            'total':  pending_punch + pending_ot + pending_sched + pending_leave,
        },
        # 本月薪資
        'salary_summary': {
            'total_count':     sal_rows['total_count'],
            'confirmed_count': sal_rows['confirmed_count'],
            'total_net':       float(sal_rows['total_net']),
            'total_allow':     float(sal_rows['total_allow']),
            'total_deduct':    float(sal_rows['total_deduct']),
        },
        # 圖表資料
        'daily_attendance':    daily_attendance,
        'leave_distribution':  leave_distribution,
        'ot_ranking':          ot_ranking,
    })


# ── Dashboard 擴充 API ────────────────────────────────────────────────────────

@bp.route('/api/dashboard/labor-cost', methods=['GET'])
@login_required
def api_dashboard_labor_cost():
    """近 12 個月人事費用趨勢"""
    from datetime import date as _dlc
    today = _dlc.today()
    months = []
    for i in range(11, -1, -1):
        m = today.month - i
        y = today.year
        while m <= 0: m += 12; y -= 1
        months.append(f'{y}-{m:02d}')
    with get_db() as conn:
        rows = conn.execute("""
            SELECT month, COALESCE(SUM(net_pay),0) as total
            FROM salary_records
            WHERE month = ANY(%s)
            GROUP BY month
        """, (months,)).fetchall()
    cost_map = {r['month']: float(r['total']) for r in rows}
    return jsonify({
        'months':     months,
        'labor_cost': [cost_map.get(m, 0) for m in months],
    })


@bp.route('/api/dashboard/attendance-heatmap', methods=['GET'])
@login_required
def api_dashboard_attendance_heatmap():
    """本月每日出勤率（熱力圖資料）"""
    from datetime import date as _dah
    import calendar as _calh
    month = request.args.get('month', '') or _dah.today().strftime('%Y-%m')
    y, mo = int(month[:4]), int(month[5:7])
    days_in = _calh.monthrange(y, mo)[1]

    with get_db() as conn:
        total_staff = conn.execute(
            "SELECT COUNT(*) as c FROM punch_staff WHERE active=TRUE"
        ).fetchone()['c']

        punch_rows = conn.execute("""
            SELECT (punched_at AT TIME ZONE 'Asia/Taipei')::date as d,
                   COUNT(DISTINCT staff_id) as cnt
            FROM punch_records
            WHERE punch_type='in'
              AND to_char(punched_at AT TIME ZONE 'Asia/Taipei','YYYY-MM')=%s
            GROUP BY d
        """, (month,)).fetchall()

        _mf = f'{y}-{mo:02d}-01'
        _ml = f'{y}-{mo:02d}-{days_in:02d}'
        leave_rows = conn.execute("""
            SELECT lr.start_date, lr.end_date, COUNT(*) as cnt
            FROM leave_requests lr
            WHERE lr.status='approved'
              AND lr.start_date <= %s AND lr.end_date >= %s
            GROUP BY lr.start_date, lr.end_date
        """, (_ml, _mf)).fetchall()

    punch_map = {str(r['d']): int(r['cnt']) for r in punch_rows}

    from datetime import date as _dah2, timedelta as _tdah
    leave_map = {}
    for lr in leave_rows:
        s = _dah2.fromisoformat(str(lr['start_date']))
        e = _dah2.fromisoformat(str(lr['end_date']))
        cur = s
        while cur <= e:
            ds = str(cur)
            if ds.startswith(month):
                leave_map[ds] = leave_map.get(ds, 0) + 1
            cur += _tdah(days=1)

    days = []
    for d in range(1, days_in + 1):
        ds = f'{y}-{mo:02d}-{d:02d}'
        cnt = punch_map.get(ds, 0)
        rate = round(cnt / total_staff, 3) if total_staff > 0 else 0
        days.append({
            'date': ds,
            'day_of_week': _dah2(y, mo, d).weekday(),
            'count': cnt,
            'attendance_rate': rate,
            'on_leave': leave_map.get(ds, 0),
        })

    return jsonify({'month': month, 'total_staff': total_staff, 'days': days})


@bp.route('/api/dashboard/leave-distribution', methods=['GET'])
@login_required
def api_dashboard_leave_distribution():
    """本年度請假類型分佈"""
    from datetime import date as _dld
    year = request.args.get('year', str(_dld.today().year))
    with get_db() as conn:
        rows = conn.execute("""
            SELECT lt.name, lt.color,
                   COUNT(*) as cnt,
                   COALESCE(SUM(lr.days), 0) as days
            FROM leave_requests lr
            JOIN leave_types lt ON lt.id=lr.leave_type_id
            WHERE lr.status='approved'
              AND EXTRACT(YEAR FROM lr.start_date)=%s
            GROUP BY lt.name, lt.color
            ORDER BY days DESC
        """, (int(year),)).fetchall()
    total = sum(float(r['days']) for r in rows)
    return jsonify({
        'year': year,
        'total_leave_days': total,
        'breakdown': [{
            'name':  r['name'],
            'color': r['color'] or '#4a7bda',
            'days':  float(r['days']),
            'count': int(r['cnt']),
            'pct':   round(float(r['days']) / total * 100, 1) if total > 0 else 0,
        } for r in rows],
    })


# ── 年度扣繳憑單 ────────────────────────────────────────────────────────────

@bp.route('/api/export/withholding', methods=['GET'])
@require_module('salary')
def api_export_withholding():
    """年度薪資所得扣繳憑單（所得類別50）"""
    from datetime import date as _dwh
    year   = request.args.get('year', str(_dwh.today().year))
    fmt    = request.args.get('format', 'html')

    fs = _get_finance_settings()
    company_name   = fs.get('company_name', '')
    company_tax_id = fs.get('company_tax_id', '')
    company_address= fs.get('company_address', '')

    with get_db() as conn:
        rows = conn.execute("""
            SELECT ps.id, ps.name, ps.national_id, ps.address,
                   COALESCE(SUM(sr.allowance_total), 0)       AS gross_salary,
                   COALESCE(SUM(sr.income_tax_withheld), 0)   AS tax_withheld,
                   COALESCE(AVG(sr.insured_salary), 0)        AS avg_insured
            FROM salary_records sr
            JOIN punch_staff ps ON ps.id = sr.staff_id
            WHERE sr.month LIKE %s AND sr.status='confirmed'
            GROUP BY ps.id, ps.name, ps.national_id, ps.address
            ORDER BY ps.name
        """, (f'{year}-%',)).fetchall()

    # 計算二代健保補充費
    def supp_nhi(gross, insured):
        base = float(gross) - float(insured) * 12
        return max(0, round(base * 0.0211, 0)) if base > 0 else 0

    data = []
    for i, r in enumerate(rows, 1):
        gross = float(r['gross_salary'])
        insured = float(r['avg_insured'])
        data.append({
            'no':          i,
            'name':        r['name'],
            'national_id': r['national_id'] or '—',
            'address':     r['address'] or '—',
            'gross':       gross,
            'supp_nhi':    supp_nhi(gross, insured),
            'tax':         float(r['tax_withheld']),
        })

    if fmt == 'xlsx':
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from io import BytesIO
        wb = wb2 = openpyxl.Workbook()
        ws = wb.active; ws.title = f'{year}年扣繳憑單'
        hfill = PatternFill('solid', fgColor='0F1C3A')
        thin  = Border(*[Side(style='thin', color='DDDDDD')]*4)
        hdrs  = ['序號','姓名','身分證字號','地址','年度薪資合計','二代健保補充費','扣繳稅額']
        ws.append(hdrs)
        for ci, h in enumerate(hdrs, 1):
            c = ws.cell(1, ci); c.font = Font(bold=True, color='FFFFFF', size=10); c.fill = hfill
            c.alignment = Alignment(horizontal='center', vertical='center'); c.border = thin
        ws.column_dimensions['A'].width = 5; ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 14; ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 16; ws.column_dimensions['F'].width = 16; ws.column_dimensions['G'].width = 12
        for d in data:
            ws.append([d['no'], d['name'], d['national_id'], d['address'],
                       d['gross'], d['supp_nhi'], d['tax']])
        buf = BytesIO(); wb.save(buf); buf.seek(0)
        from flask import Response as _FR2
        return _FR2(buf.read(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename=withholding_{year}.xlsx'})

    # HTML printable
    rows_html = ''.join(f"""
      <tr>
        <td style="text-align:center">{d['no']}</td>
        <td>{d['name']}</td>
        <td style="font-family:monospace">{d['national_id']}</td>
        <td style="font-size:11px">{d['address']}</td>
        <td style="text-align:right;font-family:monospace">{d['gross']:,.0f}</td>
        <td style="text-align:right;font-family:monospace">{d['supp_nhi']:,.0f}</td>
        <td style="text-align:right;font-family:monospace">{d['tax']:,.0f}</td>
      </tr>""" for d in data)
    html = f"""<!DOCTYPE html><html lang="zh-TW"><head>
<meta charset="UTF-8"><title>{year}年度薪資扣繳憑單</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'Noto Sans TC',sans-serif;font-size:12px;padding:20px;color:#1e2a45}}
h2{{font-size:16px;font-weight:700;margin-bottom:4px}}
.meta{{font-size:11px;color:#666;margin-bottom:16px}}
table{{width:100%;border-collapse:collapse;margin-bottom:20px}}
th{{background:#0f1c3a;color:#fff;padding:7px 10px;font-size:11px;font-weight:600;text-align:left}}
td{{padding:6px 10px;border-bottom:1px solid #eee;font-size:12px}}
tr:nth-child(even){{background:#f8f9fb}}
.note{{font-size:10px;color:#888;border-top:1px solid #ddd;padding-top:8px}}
@media print{{button{{display:none}}}}
</style></head><body>
<button onclick="window.print()" style="margin-bottom:16px;padding:6px 16px;background:#0f1c3a;color:#fff;border:none;border-radius:4px;cursor:pointer;font-size:12px">列印</button>
<h2>{year} 年度薪資所得扣繳憑單（所得類別 50）</h2>
<div class="meta">扣繳義務人：{company_name}　統一編號：{company_tax_id}　地址：{company_address}　製表日期：{_dwh.today().isoformat()}</div>
<table>
<thead><tr><th>#</th><th>員工姓名</th><th>身分證字號</th><th>地址</th><th>年度薪資合計(元)</th><th>二代健保補充費(元)</th><th>扣繳稅額(元)</th></tr></thead>
<tbody>{rows_html}</tbody>
</table>
<div class="note">※ 本報表依薪資紀錄計算，二代健保補充費 = 超出投保薪資部分 × 2.11%。扣繳稅額請依各月薪資記錄人工確認。</div>
</body></html>"""
    return html, 200, {'Content-Type': 'text/html; charset=utf-8'}


# ── 勞健保 EDI 申報 ─────────────────────────────────────────────────────────

def _get_insurance_settings():
    try:
        with get_db() as conn:
            rows = conn.execute("SELECT setting_key, setting_value FROM insurance_settings").fetchall()
        return {r['setting_key']: r['setting_value'] for r in rows}
    except Exception:
        return {}

def _roc_date(date_str):
    """Convert YYYY-MM-DD to YYYMMDD (ROC year)"""
    if not date_str: return '0000000'
    try:
        from datetime import date as _dedi
        d = _dedi.fromisoformat(str(date_str)[:10])
        return f'{d.year - 1911:03d}{d.month:02d}{d.day:02d}'
    except Exception:
        return '0000000'

def _edi_bytes(val, width, numeric=False):
    """Encode value to fixed-width bytes (Big5 for text, ASCII-padded for numeric)"""
    s = str(val or '')
    if numeric:
        return s.rjust(width, '0').encode('ascii', errors='replace')[:width]
    try:
        b = s.encode('big5', errors='replace')
    except Exception:
        b = s.encode('ascii', errors='replace')
    if len(b) < width:
        b = b + b' ' * (width - len(b))
    return b[:width]


@bp.route('/api/insurance/settings', methods=['GET'])
@require_module('salary')
def api_insurance_settings_get():
    return jsonify(_get_insurance_settings())

@bp.route('/api/insurance/settings', methods=['PUT'])
@require_module('salary')
def api_insurance_settings_put():
    b = request.get_json(force=True)
    with get_db() as conn:
        for k in ('labor_insurance_no', 'health_insurance_no', 'employer_name', 'employer_id'):
            conn.execute(
                "INSERT INTO insurance_settings VALUES (%s,%s) ON CONFLICT (setting_key) DO UPDATE SET setting_value=EXCLUDED.setting_value",
                (k, str(b.get(k, '')).strip()))
    return jsonify({'ok': True})


def _get_edi_staff(staff_ids_str):
    """Fetch staff rows for EDI, optionally filtered by comma-separated IDs."""
    with get_db() as conn:
        if staff_ids_str:
            ids = [int(x) for x in staff_ids_str.split(',') if x.strip().isdigit()]
            rows = conn.execute(
                f"SELECT * FROM punch_staff WHERE id = ANY(%s) AND active=TRUE ORDER BY name",
                (ids,)).fetchall()
        else:
            rows = conn.execute("SELECT * FROM punch_staff WHERE active=TRUE ORDER BY name").fetchall()
    return rows


@bp.route('/api/export/edi/labor-enroll', methods=['GET'])
@require_module('salary')
def api_edi_labor_enroll():
    """勞工保險加退保申報 EDI（Big5 固定寬度格式）"""
    event_type  = request.args.get('event_type', 'in')   # in=加保 out=退保
    staff_ids   = request.args.get('staff_ids', '')
    event_date  = request.args.get('event_date', '')
    cfg         = _get_insurance_settings()
    labor_no    = cfg.get('labor_insurance_no', '').ljust(8)[:8]
    event_code  = b'1' if event_type == 'in' else b'2'
    event_roc   = _roc_date(event_date).encode('ascii')

    lines = []
    for s in _get_edi_staff(staff_ids):
        gender_code = b'1' if (s.get('gender') or '').upper() in ('M', '男') else b'2'
        insured = str(int(float(s.get('insured_salary') or 0))).rjust(6, '0').encode('ascii')
        line = (
            _edi_bytes(labor_no, 8) +
            _edi_bytes(s['name'], 20) +
            _edi_bytes(s.get('national_id', ''), 10) +
            _roc_date(s.get('birth_date')).encode('ascii') +
            event_roc +
            event_code +
            insured +
            gender_code +
            b'00'   # 職業類別（一般）
        )
        lines.append(line)
    content = b'\r\n'.join(lines)
    fname   = f'labor_{"enroll" if event_type=="in" else "exit"}_{event_date or "date"}.edi'
    from flask import Response as _FRe
    return _FRe(content, mimetype='application/octet-stream',
                headers={'Content-Disposition': f'attachment; filename={fname}'})


@bp.route('/api/export/edi/labor-salary', methods=['GET'])
@require_module('salary')
def api_edi_labor_salary():
    """勞工保險投保薪資調整申報 EDI"""
    month     = request.args.get('month', '')
    staff_ids = request.args.get('staff_ids', '')
    cfg       = _get_insurance_settings()
    labor_no  = cfg.get('labor_insurance_no', '').ljust(8)[:8]
    if not month:
        from datetime import date as _dm2
        month = _dm2.today().strftime('%Y-%m')
    month_roc = f"{int(month[:4]) - 1911:03d}{month[5:7]}".encode('ascii')

    lines = []
    for s in _get_edi_staff(staff_ids):
        insured = str(int(float(s.get('insured_salary') or 0))).rjust(6, '0').encode('ascii')
        line = (
            _edi_bytes(labor_no, 8) +
            _edi_bytes(s['name'], 20) +
            _edi_bytes(s.get('national_id', ''), 10) +
            insured +
            month_roc
        )
        lines.append(line)
    content = b'\r\n'.join(lines)
    from flask import Response as _FRs
    return _FRs(content, mimetype='application/octet-stream',
                headers={'Content-Disposition': f'attachment; filename=labor_salary_{month}.edi'})


@bp.route('/api/export/edi/health-enroll', methods=['GET'])
@require_module('salary')
def api_edi_health_enroll():
    """全民健康保險加退保申報 EDI"""
    event_type = request.args.get('event_type', 'in')
    staff_ids  = request.args.get('staff_ids', '')
    event_date = request.args.get('event_date', '')
    cfg        = _get_insurance_settings()
    health_no  = cfg.get('health_insurance_no', '').ljust(10)[:10]
    event_code = b'1' if event_type == 'in' else b'2'
    event_roc  = _roc_date(event_date).encode('ascii')

    lines = []
    for s in _get_edi_staff(staff_ids):
        gender_code = b'1' if (s.get('gender') or '').upper() in ('M', '男') else b'2'
        insured = str(int(float(s.get('insured_salary') or 0))).rjust(6, '0').encode('ascii')
        line = (
            _edi_bytes(health_no, 10) +
            _edi_bytes(s['name'], 20) +
            _edi_bytes(s.get('national_id', ''), 10) +
            _roc_date(s.get('birth_date')).encode('ascii') +
            event_roc +
            event_code +
            insured +
            gender_code
        )
        lines.append(line)
    content = b'\r\n'.join(lines)
    fname   = f'health_{"enroll" if event_type=="in" else "exit"}_{event_date or "date"}.edi'
    from flask import Response as _FRh
    return _FRh(content, mimetype='application/octet-stream',
                headers={'Content-Disposition': f'attachment; filename={fname}'})


# ── 多店管理 ─────────────────────────────────────────────────────────────────

@bp.route('/api/stores', methods=['GET'])
@login_required
def api_stores_list():
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM stores ORDER BY id").fetchall()
    return jsonify([dict(r) for r in rows])

@bp.route('/api/stores', methods=['POST'])
@login_required
def api_stores_create():
    b = request.get_json(force=True)
    name = (b.get('name') or '').strip()
    code = (b.get('code') or '').strip() or None
    if not name: return jsonify({'error': '店名為必填'}), 400
    with get_db() as conn:
        row = conn.execute(
            "INSERT INTO stores (name, code, address) VALUES (%s,%s,%s) RETURNING *",
            (name, code, (b.get('address') or '').strip())
        ).fetchone()
    return jsonify(dict(row)), 201

@bp.route('/api/stores/<int:sid>', methods=['PUT'])
@login_required
def api_stores_update(sid):
    b = request.get_json(force=True)
    with get_db() as conn:
        row = conn.execute("""
            UPDATE stores SET name=%s, code=%s, address=%s, active=%s WHERE id=%s RETURNING *
        """, ((b.get('name') or '').strip(), (b.get('code') or None),
              (b.get('address') or '').strip(), bool(b.get('active', True)), sid)).fetchone()
    return jsonify(dict(row)) if row else ('', 404)

@bp.route('/api/stores/<int:sid>', methods=['DELETE'])
@login_required
def api_stores_delete(sid):
    with get_db() as conn:
        conn.execute("UPDATE punch_staff     SET store_id=NULL WHERE store_id=%s", (sid,))
        conn.execute("UPDATE punch_locations SET store_id=NULL WHERE store_id=%s", (sid,))
        conn.execute("DELETE FROM stores WHERE id=%s", (sid,))
    return jsonify({'deleted': sid})

@bp.route('/api/stores/<int:sid>/staff', methods=['GET'])
@login_required
def api_store_staff(sid):
    with get_db() as conn:
        rows = conn.execute(
            "SELECT id, name, role, active FROM punch_staff WHERE store_id=%s ORDER BY name", (sid,)
        ).fetchall()
    return jsonify([dict(r) for r in rows])

@bp.route('/api/staff/<int:sid>/store', methods=['PUT'])
@login_required
def api_staff_assign_store(sid):
    b = request.get_json(force=True)
    store_id = b.get('store_id')
    with get_db() as conn:
        conn.execute("UPDATE punch_staff SET store_id=%s WHERE id=%s", (store_id, sid))
    return jsonify({'ok': True})


# ── 排班需求 & 自動排班 ──────────────────────────────────────────────────────

@bp.route('/api/shifts/staffing-requirements', methods=['GET'])
@login_required
def api_staffing_req_get():
    with get_db() as conn:
        rows = conn.execute("""
            SELECT r.id, r.shift_type_id, r.day_of_week, r.required_count,
                   st.name as shift_name, st.color as shift_color
            FROM shift_staffing_requirements r
            JOIN shift_types st ON st.id=r.shift_type_id
            ORDER BY st.sort_order, r.day_of_week
        """).fetchall()
    return jsonify([dict(r) for r in rows])


@bp.route('/api/shifts/staffing-requirements', methods=['PUT'])
@login_required
def api_staffing_req_put():
    items = request.get_json(force=True)
    if not isinstance(items, list):
        return jsonify({'error': '格式錯誤'}), 400
    count = 0
    with get_db() as conn:
        for it in items:
            stid = int(it.get('shift_type_id', 0))
            dow  = int(it.get('day_of_week', 0))
            req  = max(0, int(it.get('required_count', 1)))
            if req == 0:
                conn.execute(
                    "DELETE FROM shift_staffing_requirements WHERE shift_type_id=%s AND day_of_week=%s",
                    (stid, dow))
            else:
                conn.execute("""
                    INSERT INTO shift_staffing_requirements (shift_type_id, day_of_week, required_count, updated_at)
                    VALUES (%s,%s,%s,NOW())
                    ON CONFLICT (shift_type_id, day_of_week)
                    DO UPDATE SET required_count=EXCLUDED.required_count, updated_at=NOW()
                """, (stid, dow, req))
            count += 1
    return jsonify({'ok': True, 'upserted': count})


@bp.route('/api/schedule/auto-generate', methods=['POST'])
@login_required
def api_auto_generate_schedule():
    """自動排班引擎：依人力需求與員工可用性生成班表建議"""
    from datetime import date as _dag, timedelta as _tdag
    import calendar as _calag

    b        = request.get_json(force=True)
    month    = (b.get('month') or '').strip()
    overwrite = bool(b.get('overwrite', False))
    if not month:
        month = _dag.today().strftime('%Y-%m')
    try:
        y, mo = int(month[:4]), int(month[5:7])
    except Exception:
        return jsonify({'error': '月份格式錯誤'}), 400

    days_in   = _calag.monthrange(y, mo)[1]
    all_dates = [_dag(y, mo, d) for d in range(1, days_in + 1)]

    with get_db() as conn:
        shift_types = conn.execute(
            "SELECT * FROM shift_types WHERE active=TRUE ORDER BY sort_order"
        ).fetchall()
        requirements = conn.execute("""
            SELECT shift_type_id, day_of_week, required_count
            FROM shift_staffing_requirements
        """).fetchall()
        staff_list = conn.execute(
            "SELECT id, name FROM punch_staff WHERE active=TRUE ORDER BY name"
        ).fetchall()

        # 本月已核准休假日期（per staff）
        leave_rows = conn.execute("""
            SELECT staff_id, start_date, end_date
            FROM leave_requests
            WHERE status='approved'
              AND start_date <= %s AND end_date >= %s
        """, (f'{y}-{mo:02d}-{days_in:02d}', f'{y}-{mo:02d}-01')).fetchall()

        # 已核准排休
        sched_rows = conn.execute("""
            SELECT staff_id, dates
            FROM schedule_requests
            WHERE status='approved'
              AND month=%s
        """, (month,)).fetchall()

        # 現有班表
        existing = conn.execute("""
            SELECT staff_id, shift_date FROM shift_assignments
            WHERE TO_CHAR(shift_date,'YYYY-MM')=%s
        """, (month,)).fetchall()

    # 建立不可上班日 set: {(staff_id, date_str)}
    off_days = set()
    for lr in leave_rows:
        s = _dag.fromisoformat(str(lr['start_date']))
        e = _dag.fromisoformat(str(lr['end_date']))
        cur = s
        while cur <= e:
            off_days.add((lr['staff_id'], str(cur)))
            cur += _tdag(days=1)
    for sr in sched_rows:
        rdates = sr['dates']
        if isinstance(rdates, str):
            try: rdates = _json.loads(rdates)
            except: rdates = []
        for ds in (rdates or []):
            off_days.add((sr['staff_id'], ds))

    # 已有班表 set（不 overwrite 時跳過）
    existing_set = {(r['staff_id'], str(r['shift_date'])) for r in existing}

    # 需求 map: {(shift_type_id, day_of_week): required_count}
    req_map = {(r['shift_type_id'], r['day_of_week']): r['required_count'] for r in requirements}

    # 排班計數器（避免連續超時）
    assigned_days  = {s['id']: [] for s in staff_list}  # staff_id -> [date]
    assignments    = []
    conflicts      = []
    staff_ids      = [s['id'] for s in staff_list]
    staff_name_map = {s['id']: s['name'] for s in staff_list}

    for date in all_dates:
        dow = date.weekday()  # 0=Mon, 6=Sun
        ds  = str(date)

        for st in shift_types:
            stid     = st['id']
            needed   = req_map.get((stid, dow), 0)
            if needed <= 0:
                continue

            # 可用員工：未請假、未排休
            available = [
                sid for sid in staff_ids
                if (sid, ds) not in off_days
            ]

            # 排除已被指派在其他班（同日）
            already_today = {a['staff_id'] for a in assignments if a['shift_date'] == ds}
            available = [sid for sid in available if sid not in already_today]

            # 排除連續 7 天（含本日）的員工
            def consecutive_days(sid, d):
                days = sorted(assigned_days[sid])
                streak = 0
                check = d
                while check in days:
                    streak += 1
                    check = str(_dag.fromisoformat(check) - _tdag(days=1))
                return streak

            available_ok = [sid for sid in available if consecutive_days(sid, ds) < 6]

            # 按本月已排天數升序（均衡分配）
            available_ok.sort(key=lambda sid: len(assigned_days[sid]))

            assigned_count = 0
            for sid in available_ok:
                if assigned_count >= needed:
                    break
                if not overwrite and (sid, ds) in existing_set:
                    assigned_count += 1
                    continue
                assignments.append({
                    'staff_id':     sid,
                    'staff_name':   staff_name_map[sid],
                    'shift_type_id': stid,
                    'shift_name':   st['name'],
                    'shift_date':   ds,
                })
                assigned_days[sid].append(ds)
                assigned_count += 1

            if assigned_count < needed:
                conflicts.append({
                    'type':   'understaffed',
                    'date':   ds,
                    'shift':  st['name'],
                    'detail': f'{ds} {st["name"]} 需要 {needed} 人，僅能排 {assigned_count} 人',
                })

    # 寫入資料庫
    inserted = 0
    if assignments:
        with get_db() as conn:
            for a in assignments:
                try:
                    if overwrite:
                        conn.execute("""
                            INSERT INTO shift_assignments (staff_id, shift_type_id, shift_date)
                            VALUES (%s,%s,%s)
                            ON CONFLICT (staff_id, shift_date) DO UPDATE
                            SET shift_type_id=EXCLUDED.shift_type_id
                        """, (a['staff_id'], a['shift_type_id'], a['shift_date']))
                    else:
                        conn.execute("""
                            INSERT INTO shift_assignments (staff_id, shift_type_id, shift_date)
                            VALUES (%s,%s,%s)
                            ON CONFLICT DO NOTHING
                        """, (a['staff_id'], a['shift_type_id'], a['shift_date']))
                    inserted += 1
                except Exception:
                    pass

    return jsonify({
        'ok':          True,
        'month':       month,
        'assignments': assignments,
        'conflicts':   conflicts,
        'summary': {
            'assigned':       inserted,
            'conflict_count': len(conflicts),
        },
    })


@bp.route('/api/attendance/anomalies', methods=['GET'])
@login_required
def api_attendance_anomalies():
    """
    偵測出勤異常：
    - 忘記打下班卡（有上班無下班）
    - 只有下班無上班
    - 遲到（上班時間晚於班別開始時間）
    """
    from datetime import date as _da, datetime as _dta, timezone as _tz, timedelta as _td
    TW    = _tz(_td(hours=8))
    today = _dta.now(TW).date()
    # Check last 7 days
    date_from = today - _td(days=7)

    with get_db() as conn:
        # 取得最近7天打卡記錄（按人、按天）
        rows = conn.execute("""
            SELECT ps.id as staff_id, ps.name, ps.role, ps.department,
                   (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date as work_date,
                   array_agg(pr.punch_type ORDER BY pr.punched_at) as types,
                   MIN(CASE WHEN pr.punch_type='in'  THEN to_char(pr.punched_at AT TIME ZONE 'Asia/Taipei','HH24:MI') END) as first_in,
                   MAX(CASE WHEN pr.punch_type='out' THEN to_char(pr.punched_at AT TIME ZONE 'Asia/Taipei','HH24:MI') END) as last_out
            FROM punch_records pr
            JOIN punch_staff ps ON ps.id = pr.staff_id
            WHERE (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date BETWEEN %s AND %s
              AND ps.active = TRUE
            GROUP BY ps.id, ps.name, ps.role, ps.department,
                     (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date
            ORDER BY work_date DESC, ps.name
        """, (date_from, today)).fetchall()

        # 取得班別指派（用於遲到／早退判斷）
        shift_rows = conn.execute("""
            SELECT sa.staff_id, sa.date, st.start_time, st.end_time, st.name as shift_name
            FROM shift_assignments sa
            JOIN shift_types st ON st.id = sa.shift_type_id
            WHERE sa.date BETWEEN %s AND %s
        """, (date_from, today)).fetchall()
        shift_map = {(r['staff_id'], str(r['date'])): r for r in shift_rows}

        # 今日應出勤但未出勤（排除請假）
        all_staff = conn.execute(
            "SELECT id, name, role, department FROM punch_staff WHERE active=TRUE"
        ).fetchall()
        today_punched_ids = {r['staff_id'] for r in rows if str(r['work_date']) == str(today)}
        on_leave_today_ids = set()
        leave_today = conn.execute("""
            SELECT DISTINCT staff_id FROM leave_requests
            WHERE status='approved' AND start_date <= %s AND end_date >= %s
        """, (today, today)).fetchall()
        for r in leave_today:
            on_leave_today_ids.add(r['staff_id'])

    anomalies = []

    # 1. 近7天：有上班但無下班卡
    for r in rows:
        types = list(r['types']) if r['types'] else []
        has_in  = 'in'  in types
        has_out = 'out' in types
        ds = str(r['work_date'])

        if has_in and not has_out and ds != str(today):
            # 昨天或更早沒打下班卡（今天的可能還沒下班）
            anomalies.append({
                'type':       'missing_out',
                'label':      '忘記下班打卡',
                'severity':   'warning',
                'staff_id':   r['staff_id'],
                'name':       r['name'],
                'role':       r['role'] or '',
                'department': r['department'] or '',
                'date':       ds,
                'detail':     f"上班 {r['first_in']}，無下班記錄",
            })

        if not has_in and has_out:
            anomalies.append({
                'type':       'missing_in',
                'label':      '忘記上班打卡',
                'severity':   'warning',
                'staff_id':   r['staff_id'],
                'name':       r['name'],
                'role':       r['role'] or '',
                'department': r['department'] or '',
                'date':       ds,
                'detail':     f"下班 {r['last_out']}，無上班記錄",
            })

        # 遲到判斷（有班別指派）
        if has_in and r['first_in']:
            shift = shift_map.get((r['staff_id'], ds))
            if shift and shift['start_time']:
                try:
                    sh, sm = map(int, str(shift['start_time'])[:5].split(':'))
                    ih, im = map(int, r['first_in'].split(':'))
                    late_mins = (ih * 60 + im) - (sh * 60 + sm)
                    if late_mins > 10:  # 超過10分鐘算遲到
                        anomalies.append({
                            'type':       'late',
                            'label':      '遲到',
                            'severity':   'warning',
                            'staff_id':   r['staff_id'],
                            'name':       r['name'],
                            'role':       r['role'] or '',
                            'department': r['department'] or '',
                            'date':       ds,
                            'detail':     f"應 {shift['start_time'][:5]} 上班，實際 {r['first_in']}（晚 {late_mins} 分鐘）",
                        })
                except Exception:
                    pass

        # 早退判斷（有班別指派）
        if has_out and r['last_out'] and ds != str(today):
            shift = shift_map.get((r['staff_id'], ds))
            if shift and shift['end_time']:
                try:
                    eh, em = map(int, str(shift['end_time'])[:5].split(':'))
                    oh, om = map(int, r['last_out'].split(':'))
                    early_mins = (eh * 60 + em) - (oh * 60 + om)
                    if early_mins > 15:  # 超過15分鐘算早退
                        anomalies.append({
                            'type':       'early',
                            'label':      '早退',
                            'severity':   'warning',
                            'staff_id':   r['staff_id'],
                            'name':       r['name'],
                            'role':       r['role'] or '',
                            'department': r['department'] or '',
                            'date':       ds,
                            'detail':     f"應 {shift['end_time'][:5]} 下班，實際 {r['last_out']}（早 {early_mins} 分鐘）",
                        })
                except Exception:
                    pass

    # 2. 今日未出勤（不含請假）
    for s in all_staff:
        if s['id'] not in today_punched_ids and s['id'] not in on_leave_today_ids:
            anomalies.append({
                'type':       'absent',
                'label':      '今日未出勤',
                'severity':   'error',
                'staff_id':   s['id'],
                'name':       s['name'],
                'role':       s['role'] or '',
                'department': s['department'] or '',
                'date':       str(today),
                'detail':     '今日尚無打卡記錄且未請假',
            })

    # Sort: error > warning > info, then by date desc
    sev_order = {'error': 0, 'warning': 1, 'info': 2}
    anomalies.sort(key=lambda x: (sev_order.get(x['severity'], 9), x['date']))
    return jsonify({'anomalies': anomalies, 'count': len(anomalies), 'checked_from': str(date_from)})


# ═══════════════════════════════════════════════════════════════════
# Feature: Staff Termination (離職流程)
# ═══════════════════════════════════════════════════════════════════
